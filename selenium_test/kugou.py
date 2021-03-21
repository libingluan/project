# 读取一个文件有100万行，获取相同字段内容和次数
from openpyxl import load_workbook

wb = load_workbook("./case.xlsx")
sheet1 = wb['Sheet1']
rows = sheet1.max_row  # 获取行数
list=[]
for i in range(rows+1):
    val = sheet1.cell(i, 1).value
    list = list.append(val)  # 获取每一行的内容，生成列表
length = len(list) # 求长度
j = 0
while j < length:
    time = 1
    p = list[j]
    while j+1 <= length-1 and list[j] == list[j+1]:
        time += 1
        j += 1
    print(p, time)
    j+1





