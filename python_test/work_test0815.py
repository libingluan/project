
# Python群发一个excel表格中的所有联系人的邮件
# 1.多个sheet表格依次发送所有的邮件
# 2.每个邮件的内容不一样
from email.header import Header

import openpyxl
import smtplib
from email.mime.text import MIMEText
# list = [["657367560@qq.com", "测试内容11"], ["miya657367560@163.com", "测试内容222"], ["miya1114@126.com", "测试内容333"]]
# # 创建一个文件
# book = openpyxl.workbook
# sheet1 = book["Sheet1"]
# for i in list:
#     sheet1.append(i)
# book.save("email_info.xlsx")

book = openpyxl.load_workbook("email_info.xlsx")
sheet1 = book.active

rows = sheet1.max_row
cols = sheet1.max_column
print(rows)
email_lis = []
content_lis = []
for i in range(2, rows+1):
    emails = sheet1.cell(i, 1).value
    email_lis.append(sheet1.cell(i, 1).value)
    content_lis.append(sheet1.cell(i, 2).value)
print(email_lis, content_lis)

class email():
    def __init__(self, sender, content,password, receiver, smtpserver):
        self.sender = sender
        self.content = content
        self.password = password
        self.receiver = receiver
        self.smtpserver = smtpserver

    def send_email(self):
        '''
        发送邮件，不带附件信息
        :return:
        '''
        # 第一步：创建数据
        # sender = 'miya657367560@163.com'
        # receiver = 'miya657367560@163.com'
        # password = 'CGGGRNLSEZXYKJVL'
        # smtpserver = 'smtp.163.com'
        # 第二步：构建邮件的内容：MIMEText
        content = MIMEText(self.content, 'plain', 'utf-8')
        content['From'] = Header(self.sender, 'utf-8')
        content['To'] = self.receiver
        content['From'] = Header('学习邮件发送主题')
        # 第三步：发送邮件 smtplib
        serve = smtplib.SMTP(self.smtpserver, 25)  # 定义邮件服务器
        serve.set_debuglevel(1)  # 打印日志
        serve.login(self.sender, self.password)
        serve.sendmail(self.sender, self.receiver, self.content.as_string())
        serve.quit()


    def attachment_email(self):
         '''
         发送邮件带附件信息
         :return:
         '''

