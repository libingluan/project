# 一、xlrd和xlwt读取文件练习题：
import xlwt
import xlrd
import openpyxl
# 1：用xlwt在一个.xls文件中写入数据（数据自定义），并保存为demo.xls
book = xlwt.Workbook()
sheet1 = book.add_sheet("sheet1")
sheet1.write(0, 0, "姓名")
sheet1.write(0, 1, "性别")
sheet1.write(0, 2, "年龄")
row2 = ["xiaomi", "女", "20"]
for i in range(len(row2)):
    sheet1.write(1, i, row2[i])
book.save("demo.xls")
# 2：用xlrd读取.xls文件(文件自行准备)中的内容，要求如下：
# 	利用for循环，按行读取的方式，读取.xls文件的所有内容，并在控制台输出
book1 = xlrd.open_workbook("demo.xls")
sh1 = book1.sheet_by_index(0)
nrows = sh1.nrows   # 获取总行数
ncols = sh1.ncols    # 获取总列数
for i in range(nrows):
    row = sh1.row_values(i)
    print(row)
# 	利用for循环，按列读取的方式，读取.xls文件的所有内容，并在控制台输出
for j in range(ncols):
    col = sh1.col_values(j)
    print(col)
# 	利用for循环，按单元格读取的方式，读取.xls文件的所有内容，并在控制台输出
for i in range(nrows):
    for j in range(ncols):
        cell = sh1.cell_value(i, j)
        print(cell)
# 3：用xlrd读取.xlsx文件(文件自行准备)中的内容，要求如下：
# 	利用for循环，按行读取的方式，读取.xlsx文件的所有内容，并在控制台输出
book2 = xlrd.open_workbook("data.xlsx")
sh1 = book2.sheet_by_index(0)
nrows2 = sh1.nrows
ncols2 = sh1.ncols
for i in range(nrows2):
    row = sh1.row_values(i)
    print(row)
# 	利用for循环，按列读取的方式，读取.xlsx文件的所有内容，并在控制台输出
for j in range(ncols2):
    col = sh1.col_values(j)
    print(col)
# 	利用for循环，按单元格读取的方式，读取.xlsx文件的所有内容，并在控制台输出
for i in range(nrows2):
    for j in range(ncols2):
        cell = sh1.cell(i, j).value
        print(cell)
# 二、openpyxl练习题：
# 1：使用openpyxl创建一个新的.xlsx文件，把如下信息写入文件中，并保存为userinfo.xlsx：
name2Age = [['张飞',  38], ['赵云',  27], ['许褚',  36], ['典韦',  38],
            ['关羽',  39], ['黄忠',  49], ['徐晃',  43], ['马超',  23]]
#book = openpyxl.load_workbook("userinfo.xlsx")
book = openpyxl.workbook()  #新建一个文件
sheet1 = book["Sheet1"]
#sheet1 = book.active
print(len(name2Age))
# for i in range(len(name2Age)):
#     sheet1.append(name2Age[i])
for i in name2Age:
    sheet1.append(i)
book.save("userinfo.xlsx")
# 2：用openpyxl读取.xlsx文件(文件自行准备)中的内容，要求如下：
#      利用for循环，按行读取的方式，读取.xlsx文件的所有内容，并在控制台输出，使用sh1.cell(1,2).value的方式读取
# for row in sheet1.iter_rows(max_row=len(name2Age)):
#     for cell in row:
#         print(cell.value, end="\t")
#     print()
rows = sheet1.max_row
cols = sheet1.max_column
print(rows, cols)
for i in range(1, rows+1):
    for j in range(1, cols+1):
        print(sheet1.cell(i, j).value)
#      利用for循环，按列读取的方式，读取.xlsx文件的所有内容，并在控制台输出，使用sh1.cell(1,2).value的方式读取
for j in range(1, cols+1):
    for i in range(1, rows+1):
        print(sheet1.cell(i, j).value)
