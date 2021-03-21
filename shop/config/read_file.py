#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time : 2021-03-20 23:38
# @Author : libingluan
# @File : read_file.py
# @Software: PyCharm
import yaml

# with open('login.yaml','r')as file:
#     data = yaml.load(file)
#     print(data)
#
#
# f = open("login.xlsx",'r')
# msg =f.readlines()
# print(msg)


import openpyxl
book = openpyxl.load_workbook("login.xlsx","r")
sh1 = book.get_sheet_by_name("Sheet1")
#sh1 = book.active
rows=sh1.max_row
columns = sh1.max_column
print("行数",rows)
print("列数",columns)
lists = []
for row in sh1.iter_rows(min_row=2,max_row=rows):
    rowlist= [] #定义一个行的空列表
    for cell in row:
        print(cell.value,end="\t")
        rowlist.append(cell.value) # 将每一行的内容组成一个列表
    lists.append(rowlist)  #将所有的列表在组成一个新的列表
print(lists)



# import xlrd
# book = xlrd.open_workbook("login.xlsx",'r')
# sh1 = book.sheet_by_index(0)
#
# rows = sh1.nrows  #获取行数
# data=[]
# for i in range(1,rows):
#     print(sh1.row_values(i))  #获取的行数是一个列表
#     data.append(sh1.row_values(i))
# print(data)